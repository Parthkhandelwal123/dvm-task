from openpyxl import load_workbook


class Course:
    def __init__(self, course_code, course_name, exam_date):
        self.course_code = course_code
        self.course_name = course_name
        self.exam_date = exam_date
        self.sections = []

    def get_all_sections(self):
        return self.sections

    def __str__(self):
        return f"Course Code: {self.course_code}\nCourse Name: {self.course_name}\nExam Date: {self.exam_date}"

    def populate_section(self, section_id, day_slots):
        new_section = Section(section_id, day_slots)
        self.sections.append(new_section)
        print(f"Section {section_id} added successfully for {self.course_code}")


class Section:
    def __init__(self, section_id, day_slots):
        self.section_id = section_id
        self.day_slots = day_slots  # this is a dictionary type

    def get_section_info(self):
        return f"Section ID: {self.section_id}, Day Slots: {self.day_slots}"


import csv

class Timetable:
    def __init__(self):
        self.courses = {}

    def enroll_subject(self, course):
        if course.course_code in self.courses:
            print(f"Error: Course {course.course_code} is already enrolled.")
        else:
            self.courses[course.course_code] = course
            print(f"{course.course_code} enrolled successfully.")

    def check_clashes(self):
        # Implement your logic to check for clashes between sections or exams here
        pass

    def export_to_csv(self, filename):
        with open(filename, 'w', newline='') as csvfile:
            fieldnames = ['Course Code', 'Course Name', 'Exam Date', 'Sections']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for course in self.courses.values():
                writer.writerow({
                    'Course Code': course.course_code,
                    'Course Name': course.course_name,
                    'Exam Date': course.exam_date,
                    'Sections': ';'.join([f"{section.section_id}:{section.day_slots}" for section in course.sections])
                })

        print(f"Timetable exported to {filename} successfully.")



def populate_courses_from_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    courses = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        course_code, course_name, exam_date, sections_data = row[:4]
        
        if not sections_data:
            # Handle the case where 'Sections' column is empty
            sections = []
        else:
            sections = sections_data.split(',')
        
        existing_course = next((course for course in courses if course.course_code == course_code), None)

        if existing_course:
            for section_info in sections:
                section_id, day_slots = section_info.split(':')
                existing_course.populate_section(section_id.strip(), day_slots.strip())
        else:
            new_course = Course(course_code, course_name, exam_date)
            for section_info in sections:
                section_id, day_slots = section_info.split(':')
                new_course.populate_section(section_id.strip(), day_slots.strip())
            courses.append(new_course)

    return courses



def main():
    timetable = Timetable()

    # Populate courses from an Excel spreadsheet
    courses = populate_courses_from_excel("courses.xlsx")

    # Enroll subjects to the timetable
    for course in courses:
        timetable.enroll_subject(course)

    # Check for clashes
    timetable.check_clashes()

    # Export timetable to a CSV file
    timetable.export_to_csv("timetable.csv")


if __name__ == "__main__":
    main()
